import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
from src.schemas import PrescriptionData, LineItem

class ExcelGenerator:
    """Service for generating Excel files from structured prescription data"""
    
    def __init__(self, output_dir: str = "data/output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_excel(self, prescription_data: PrescriptionData, document_id: str) -> str:
        """Generate Excel file from prescription data"""
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{document_id}_processed_{timestamp}.xlsx"
        output_path = self.output_dir / filename
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Line Items (main data)
            line_items_df = self._create_line_items_dataframe(prescription_data)
            line_items_df.to_excel(writer, sheet_name='Line_Items', index=False)
            
            # Sheet 2: Prescription Summary
            summary_df = self._create_summary_dataframe(prescription_data)
            summary_df.to_excel(writer, sheet_name='Prescription_Summary', index=False)
            
            # Sheet 3: Clinical Codes
            codes_df = self._create_codes_dataframe(prescription_data)
            codes_df.to_excel(writer, sheet_name='Clinical_Codes', index=False)
            
            # Format the sheets
            self._format_excel_sheets(writer, line_items_df, summary_df, codes_df)
        
        print(f"Excel file generated: {output_path}")
        return str(output_path)
    
    def _create_line_items_dataframe(self, prescription_data: PrescriptionData) -> pd.DataFrame:
        """Create dataframe for line items sheet"""
        
        rows = []
        for i, item in enumerate(prescription_data.line_items, 1):
            row = {
                'Item_No': i,
                'Medication_Name': item.item_name or '',
                'Quantity': item.quantity or '',
                'Unit_Price': item.unit_price,
                'Total_Price': item.total_price,
                'Dosage': item.dosage or '',
                'Frequency': item.frequency or '',
                'Duration': item.duration or '',
                'Clinical_Code': item.clinical_code or '',
                'Code_Type': item.code_type or '',
                'Raw_Text': item.raw_text or ''
            }
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_summary_dataframe(self, prescription_data: PrescriptionData) -> pd.DataFrame:
        """Create dataframe for prescription summary sheet"""
        
        summary_data = [
            ['Document ID', prescription_data.document_id],
            ['Patient Name', prescription_data.patient_name or 'N/A'],
            ['Patient ID', prescription_data.patient_id or 'N/A'],
            ['Doctor Name', prescription_data.doctor_name or 'N/A'],
            ['Doctor ID', prescription_data.doctor_id or 'N/A'],
            ['Prescription Date', prescription_data.prescription_date or 'N/A'],
            ['Pharmacy Name', prescription_data.pharmacy_name or 'N/A'],
            ['Total Amount', prescription_data.total_amount or 'N/A'],
            ['Number of Line Items', len(prescription_data.line_items)],
            ['Items with Clinical Codes', sum(1 for item in prescription_data.line_items if item.clinical_code)],
            ['Processing Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        return pd.DataFrame(summary_data, columns=['Field', 'Value'])
    
    def _create_codes_dataframe(self, prescription_data: PrescriptionData) -> pd.DataFrame:
        """Create dataframe for clinical codes sheet"""
        
        codes_data = []
        for i, item in enumerate(prescription_data.line_items, 1):
            if item.clinical_code:
                codes_data.append({
                    'Item_No': i,
                    'Medication_Name': item.item_name,
                    'Clinical_Code': item.clinical_code,
                    'Code_Type': item.code_type,
                    'Code_Description': self._get_code_description(item.clinical_code, item.code_type)
                })
        
        if not codes_data:
            # Add empty row if no codes found
            codes_data.append({
                'Item_No': 'N/A',
                'Medication_Name': 'No clinical codes assigned',
                'Clinical_Code': 'N/A',
                'Code_Type': 'N/A',
                'Code_Description': 'N/A'
            })
        
        return pd.DataFrame(codes_data)
    
    def _get_code_description(self, code: str, code_type: str) -> str:
        """Get description for clinical code (placeholder implementation)"""
        # In a real implementation, this would lookup the code in a clinical database
        if code_type == 'NDC':
            return f"NDC Code: {code}"
        elif code_type == 'ICD-10':
            return f"ICD-10 Code: {code}"
        elif code_type == 'CPT':
            return f"CPT Code: {code}"
        elif code_type == 'HCPCS':
            return f"HCPCS Code: {code}"
        else:
            return f"Clinical Code: {code}"
    
    def _format_excel_sheets(self, writer, line_items_df: pd.DataFrame, 
                           summary_df: pd.DataFrame, codes_df: pd.DataFrame):
        """Format Excel sheets for better readability"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Format Line Items sheet
            ws_items = writer.sheets['Line_Items']
            self._format_sheet_headers(ws_items)
            self._adjust_column_widths(ws_items, line_items_df)
            
            # Format Summary sheet
            ws_summary = writer.sheets['Prescription_Summary']
            self._format_sheet_headers(ws_summary)
            self._adjust_column_widths(ws_summary, summary_df)
            
            # Format Codes sheet
            ws_codes = writer.sheets['Clinical_Codes']
            self._format_sheet_headers(ws_codes)
            self._adjust_column_widths(ws_codes, codes_df)
            
        except ImportError:
            print("openpyxl styling not available, using basic formatting")
    
    def _format_sheet_headers(self, worksheet):
        """Format sheet headers with bold font and background color"""
        try:
            from openpyxl.styles import Font, PatternFill
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
        except:
            pass
    
    def _adjust_column_widths(self, worksheet, df: pd.DataFrame):
        """Adjust column widths based on content"""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
        except:
            pass
    
    def generate_batch_summary(self, results: List[Dict[str, Any]]) -> str:
        """Generate summary Excel for batch processing results"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"batch_processing_summary_{timestamp}.xlsx"
        output_path = self.output_dir / filename
        
        # Create summary data
        summary_data = []
        for result in results:
            summary_data.append({
                'Document_ID': result.get('document_id', 'N/A'),
                'Status': 'Success' if result.get('success', False) else 'Failed',
                'Output_File': Path(result.get('output_file', '')).name if result.get('output_file') else 'N/A',
                'Error_Count': len(result.get('errors', [])),
                'Errors': '; '.join(result.get('errors', [])) if result.get('errors') else 'None'
            })
        
        summary_df = pd.DataFrame(summary_data)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Batch_Summary', index=False)
            
            # Add statistics
            stats_data = [
                ['Total Files Processed', len(results)],
                ['Successful', sum(1 for r in results if r.get('success', False))],
                ['Failed', sum(1 for r in results if not r.get('success', False))],
                ['Success Rate', f"{sum(1 for r in results if r.get('success', False)) / len(results) * 100:.1f}%" if results else "0%"],
                ['Processing Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            stats_df = pd.DataFrame(stats_data, columns=['Metric', 'Value'])
            stats_df.to_excel(writer, sheet_name='Statistics', index=False)
        
        return str(output_path)